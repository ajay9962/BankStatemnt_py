# app.py
from flask import Flask, render_template, request, redirect, url_for, send_file, session, jsonify
import pdfplumber
import pandas as pd
import os, tempfile, re, uuid
from datetime import datetime
import json
import io
from werkzeug.utils import secure_filename
from openpyxl.styles import Font  # Add this import
import openpyxl.styles # Added for PatternFill

app = Flask(__name__)
app.secret_key = 'mysecretkey'

# In-memory storage for session data
temp_storage = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    try:
        file = request.files['pdf']
        password = request.form.get('password', '')
        session_id = str(uuid.uuid4())
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(filepath)

        file_ext = os.path.splitext(file.filename)[1].lower()
        rows = []
        current_row = None
        
        if file_ext in ['.pdf']:
            # Try to open PDF without password first
            try:
                with pdfplumber.open(filepath) as pdf:
                    password_needed = False
            except Exception as e:
                password_needed = True
            if password_needed:
                if not password:
                    return render_template('error.html', error='This PDF is password-protected. Please provide the password.')
                try:
                    with pdfplumber.open(filepath, password=password) as pdf:
                        pdf_pages = pdf.pages
                except Exception as e:
                    return render_template('error.html', error='Incorrect password or unable to open PDF: ' + str(e))
                # Now process as before, but with password
                with pdfplumber.open(filepath, password=password) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        if not text:
                            continue
                        lines = text.split('\n')
                        for line in lines:
                            line = line.strip()
                            header_keywords = ["date", "details", "ref", "debit", "credit", "balance"]
                            if not line or any(x in line.lower() for x in [
                                "please do not share", "computer generated", "bank never ask", 
                                "statement", "account", "opening balance", "closing balance",
                                "page", "continue", "continued"
                            ]) or all(h in line.lower() for h in header_keywords):
                                continue
                            patterns = [
                                r'^(\d{2}\s+\w+\s+\d{4})\s+(.*?)\s+(\w+\d+|\d+)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$',
                                r'^(\d{2}\s+\w+\s+\d{4})\s+(.*?)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$',
                                r'^(\d{2}/\d{2}/\d{4})\s+(.*?)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$',
                                r'^(\d{2}-\d{2}-\d{4})\s+(.*?)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$'
                            ]
                            match = None
                            for pattern in patterns:
                                match = re.match(pattern, line)
                                if match:
                                    break
                            if match:
                                groups = match.groups()
                                if len(groups) == 6:
                                    date = groups[0]
                                    details = groups[1].strip()
                                    ref_no = groups[2]
                                    debit = groups[3].replace(',', '') if groups[3] != '-' else ''
                                    credit = groups[4].replace(',', '') if groups[4] and groups[4] != '-' else ''
                                    balance = groups[5].replace(',', '')
                                else:
                                    date = groups[0]
                                    details = groups[1].strip()
                                    ref_no = ''
                                    debit = groups[2].replace(',', '') if groups[2] != '-' else ''
                                    credit = groups[3].replace(',', '') if groups[3] and groups[3] != '-' else ''
                                    balance = groups[4].replace(',', '')
                                current_row = [date, details, ref_no, debit, credit, balance, '']
                                rows.append(current_row)
                            elif current_row and line:
                                current_row[1] += ' ' + line.strip()
                df = pd.DataFrame(rows, columns=["Date", "Details", "Ref No./Cheque No", "Debit", "Credit", "Balance", "Comment"])
            else:
                # Not password protected, process as before
                with pdfplumber.open(filepath) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        text = page.extract_text()
                        if not text:
                            continue
                        lines = text.split('\n')
                        for line in lines:
                            line = line.strip()
                            header_keywords = ["date", "details", "ref", "debit", "credit", "balance"]
                            if not line or any(x in line.lower() for x in [
                                "please do not share", "computer generated", "bank never ask", 
                                "statement", "account", "opening balance", "closing balance",
                                "page", "continue", "continued"
                            ]) or all(h in line.lower() for h in header_keywords):
                                continue
                            patterns = [
                                r'^(\d{2}\s+\w+\s+\d{4})\s+(.*?)\s+(\w+\d+|\d+)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$',
                                r'^(\d{2}\s+\w+\s+\d{4})\s+(.*?)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$',
                                r'^(\d{2}/\d{2}/\d{4})\s+(.*?)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$',
                                r'^(\d{2}-\d{2}-\d{4})\s+(.*?)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)\s+(\d+\.\d{2}|[\d,]+\.\d{2}|-)?\s+(\d+\.\d{2}|[\d,]+\.\d{2})$'
                            ]
                            match = None
                            for pattern in patterns:
                                match = re.match(pattern, line)
                                if match:
                                    break
                            if match:
                                groups = match.groups()
                                if len(groups) == 6:
                                    date = groups[0]
                                    details = groups[1].strip()
                                    ref_no = groups[2]
                                    debit = groups[3].replace(',', '') if groups[3] != '-' else ''
                                    credit = groups[4].replace(',', '') if groups[4] and groups[4] != '-' else ''
                                    balance = groups[5].replace(',', '')
                                else:
                                    date = groups[0]
                                    details = groups[1].strip()
                                    ref_no = ''
                                    debit = groups[2].replace(',', '') if groups[2] != '-' else ''
                                    credit = groups[3].replace(',', '') if groups[3] and groups[3] != '-' else ''
                                    balance = groups[4].replace(',', '')
                                current_row = [date, details, ref_no, debit, credit, balance, '']
                                rows.append(current_row)
                            elif current_row and line:
                                current_row[1] += ' ' + line.strip()
                df = pd.DataFrame(rows, columns=["Date", "Details", "Ref No./Cheque No", "Debit", "Credit", "Balance", "Comment"])
        elif file_ext in ['.xlsx', '.xls']:
            df = pd.read_excel(filepath)
            # Ensure all required columns exist
            for col in ["Date", "Details", "Ref No./Cheque No", "Debit", "Credit", "Balance", "Comment"]:
                if col not in df.columns:
                    df[col] = ''
            df = df[["Date", "Details", "Ref No./Cheque No", "Debit", "Credit", "Balance", "Comment"]]
        elif file_ext in ['.csv']:
            df = pd.read_csv(filepath)
            for col in ["Date", "Details", "Ref No./Cheque No", "Debit", "Credit", "Balance", "Comment"]:
                if col not in df.columns:
                    df[col] = ''
            df = df[["Date", "Details", "Ref No./Cheque No", "Debit", "Credit", "Balance", "Comment"]]
        else:
            return render_template('error.html', error='Unsupported file type.')
        # Clean and process data
        df = clean_dataframe(df)
        # Store in session
        temp_storage[session_id] = {
            'df': df,
            'filename': file.filename,
            'processed_at': datetime.now().isoformat()
        }
        session['id'] = session_id
        # Clean up temporary file
        os.remove(filepath)
        os.rmdir(temp_dir)
        return redirect(url_for('preview'))

    except Exception as e:
        return render_template('error.html', error=str(e))

def clean_dataframe(df):
    """Clean and validate DataFrame data"""
    # Remove rows with invalid data
    df = df.dropna(subset=['Date', 'Details'])
    
    # Clean monetary values
    for col in ['Debit', 'Credit', 'Balance']:
        df[col] = df[col].astype(str).str.replace(',', '').replace('', '0')
    
    # Sort by date (attempt to parse different date formats)
    try:
        df['Date_parsed'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.sort_values('Date_parsed').drop('Date_parsed', axis=1)
    except:
        pass
    
    # Reset index
    df = df.reset_index(drop=True)
    
    return df

@app.route('/preview')
def preview():
    sid = session.get('id')
    if not sid or sid not in temp_storage:
        return redirect('/')
    
    session_data = temp_storage[sid]
    df = session_data['df']
    preview_data = df.head(10)
    
    return render_template('preview.html', 
                         data=preview_data.to_dict(orient='records'),
                         filename=session_data['filename'])

@app.route('/analysis')
def analysis():
    sid = session.get('id')
    if not sid or sid not in temp_storage:
        return redirect('/')
    
    session_data = temp_storage[sid]
    df = session_data['df']
    
    # Calculate summary statistics
    try:
        total_debit = pd.to_numeric(df["Debit"].replace('', '0'), errors='coerce').sum()
        total_credit = pd.to_numeric(df["Credit"].replace('', '0'), errors='coerce').sum()
        final_balance = pd.to_numeric(df["Balance"].replace('', '0'), errors='coerce').iloc[-1] if len(df) > 0 else 0
        
        summary = {
            'debit': total_debit,
            'credit': total_credit,
            'balance': final_balance,
            'net_flow': total_credit - total_debit,
            'transaction_count': len(df)
        }
    except Exception as e:
        summary = {
            'debit': 0,
            'credit': 0,
            'balance': 0,
            'net_flow': 0,
            'transaction_count': 0
        }
    
    return render_template('analysis.html', 
                         data=df.to_dict(orient='records'),
                         summary=summary,
                         filename=session_data['filename'])

@app.route('/update_comment', methods=['POST'])
def update_comment():
    try:
        sid = session.get('id')
        if not sid or sid not in temp_storage:
            return jsonify({'success': False, 'error': 'Session not found'})
        
        index = int(request.form.get('index'))
        comment = request.form.get('comment', '')
        
        # Update the dataframe
        df = temp_storage[sid]['df']
        if 0 <= index < len(df):
            df.loc[index, 'Comment'] = comment
            temp_storage[sid]['df'] = df
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Invalid index'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_transaction', methods=['POST'])
def delete_transaction():
    try:
        sid = session.get('id')
        if not sid or sid not in temp_storage:
            return jsonify({'success': False, 'error': 'Session not found'})
        index = int(request.form.get('index'))
        df = temp_storage[sid]['df']
        if 0 <= index < len(df):
            df = df.drop(index).reset_index(drop=True)
            temp_storage[sid]['df'] = df
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Invalid index'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_transactions', methods=['POST'])
def delete_transactions():
    try:
        sid = session.get('id')
        if not sid or sid not in temp_storage:
            return jsonify({'success': False, 'error': 'Session not found'})
        indices = request.json.get('indices', [])
        df = temp_storage[sid]['df']
        # Remove the rows by index
        df = df.drop(indices).reset_index(drop=True)
        temp_storage[sid]['df'] = df
        # Recalculate totals
        total_debit = pd.to_numeric(df["Debit"].replace('', '0'), errors='coerce').sum()
        total_credit = pd.to_numeric(df["Credit"].replace('', '0'), errors='coerce').sum()
        final_balance = pd.to_numeric(df["Balance"].replace('', '0'), errors='coerce').iloc[-1] if len(df) > 0 else 0
        summary = {
            'debit': float(total_debit),
            'credit': float(total_credit),
            'balance': float(final_balance),
            'transaction_count': int(len(df))
        }
        return jsonify({'success': True, 'summary': summary})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    try:
        sid = session.get('id')
        if not sid or sid not in temp_storage:
            return jsonify({'success': False, 'error': 'Session not found'})
        df = temp_storage[sid]['df']
        # Get fields from form
        date = request.form.get('date', '').strip()
        details = request.form.get('details', '').strip()
        ref_no = request.form.get('ref_no', '').strip()
        debit = request.form.get('debit', '').strip()
        credit = request.form.get('credit', '').strip()
        balance = request.form.get('balance', '').strip()
        comment = request.form.get('comment', '').strip()
        # Append new row
        new_row = {
            'Date': date,
            'Details': details,
            'Ref No./Cheque No': ref_no,
            'Debit': debit,
            'Credit': credit,
            'Balance': balance,
            'Comment': comment
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        temp_storage[sid]['df'] = df
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/export/xlsx')
def export_xlsx():
    sid = session.get('id')
    if not sid or sid not in temp_storage:
        return redirect('/')
    
    session_data = temp_storage[sid]
    df = session_data['df']
    
    # Get from and to dates for filename
    try:
        dates = pd.to_datetime(df['Date'], errors='coerce')
        from_date = dates.min().strftime('%Y%m%d') if not dates.isna().all() else 'unknown'
        to_date = dates.max().strftime('%Y%m%d') if not dates.isna().all() else 'unknown'
    except:
        from_date = 'unknown'
        to_date = 'unknown'
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Bank Statement', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Bank Statement']
        
        # Format headers (use openpyxl Font)
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
    
    output.seek(0)
    
    filename = f"bank_statement_{from_date}_to_{to_date}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/csv')
def export_csv():
    if 'id' not in session or session['id'] not in temp_storage:
        return redirect(url_for('index'))
    
    data = temp_storage[session['id']]['df']
    filename = temp_storage[session['id']]['filename']
    base_name = os.path.splitext(filename)[0]
    
    # Create CSV in memory
    output = io.StringIO()
    data.to_csv(output, index=False)
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'{base_name}_transactions.csv'
    )

@app.route('/get_filter_options')
def get_filter_options():
    """Get available filter options for the frontend"""
    if 'id' not in session or session['id'] not in temp_storage:
        return jsonify({'error': 'No data available'})
    
    df = temp_storage[session['id']]['df']
    
    # Simple person extraction - focus on key patterns
    persons = set()
    for detail in df['Details']:
        if pd.isna(detail) or detail == '':
            continue
            
        detail_str = str(detail).upper()
        
        # Extract UPI IDs (name@bank)
        upi_matches = re.findall(r'([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})', detail_str)
        for upi in upi_matches:
            name_part = upi.split('@')[0]
            if len(name_part) > 2:
                persons.add(name_part)
        
        # Extract from TRANSFER patterns
        if 'TRANSFER' in detail_str:
            transfer_match = re.search(r'TRANSFER\s+(?:TO|FROM)\s+([A-Z0-9\s]+?)(?:\s+\d|$)', detail_str)
            if transfer_match:
                transfer_text = transfer_match.group(1).strip()
                if len(transfer_text) > 2:
                    persons.add(transfer_text)
        
        # Extract from PAY patterns
        if 'PAY' in detail_str:
            pay_match = re.search(r'PAY\s+(?:TO\s+)?([A-Z\s]+?)(?:\s+\d|$)', detail_str)
            if pay_match:
                pay_text = pay_match.group(1).strip()
                if len(pay_text) > 2:
                    persons.add(pay_text)
        
        # Extract account numbers and look for names around them
        account_matches = re.findall(r'(\d{10,})', detail_str)
        for account in account_matches:
            # Look for text before account number
            before_parts = detail_str.split(account)[0].strip().split()
            if before_parts:
                # Take last 2-3 words before account number
                potential_name = ' '.join(before_parts[-3:])
                if len(potential_name) > 3 and not potential_name.isdigit():
                    persons.add(potential_name)
        
        # Extract merchant names before bank names
        bank_match = re.search(r'([A-Z\s]+?)\s+(?:HDFC|SBI|ICICI|AXIS|PNB|BOB|CANARA|UNION|KOTAK|YES|IDFC)', detail_str)
        if bank_match:
            bank_text = bank_match.group(1).strip()
            if len(bank_text) > 2:
                persons.add(bank_text)
    
    # Get amount ranges
    debit_amounts = df['Debit'].replace('', '0').astype(float)
    credit_amounts = df['Credit'].replace('', '0').astype(float)
    
    debit_ranges = [
        {'label': '₹0 - ₹1,000', 'min': 0, 'max': 1000},
        {'label': '₹1,000 - ₹5,000', 'min': 1000, 'max': 5000},
        {'label': '₹5,000 - ₹10,000', 'min': 5000, 'max': 10000},
        {'label': '₹10,000 - ₹50,000', 'min': 10000, 'max': 50000},
        {'label': '₹50,000+', 'min': 50000, 'max': float('inf')}
    ]
    
    credit_ranges = [
        {'label': '₹0 - ₹1,000', 'min': 0, 'max': 1000},
        {'label': '₹1,000 - ₹5,000', 'min': 1000, 'max': 5000},
        {'label': '₹5,000 - ₹10,000', 'min': 5000, 'max': 10000},
        {'label': '₹10,000 - ₹50,000', 'min': 10000, 'max': 50000},
        {'label': '₹50,000+', 'min': 50000, 'max': float('inf')}
    ]
    
    return jsonify({
        'persons': sorted(list(persons)),
        'debit_ranges': debit_ranges,
        'credit_ranges': credit_ranges
    })

@app.route('/debug_transactions')
def debug_transactions():
    """Debug route to see transaction details"""
    if 'id' not in session or session['id'] not in temp_storage:
        return jsonify({'error': 'No data available'})
    
    df = temp_storage[session['id']]['df']
    
    # Show first 10 transaction details
    sample_details = []
    for idx, row in df.head(10).iterrows():
        sample_details.append({
            'index': idx,
            'date': row['Date'],
            'details': row['Details'],
            'debit': row['Debit'],
            'credit': row['Credit']
        })
    
    return jsonify({
        'total_transactions': len(df),
        'sample_details': sample_details
    })

@app.route('/debug_persons')
def debug_persons():
    """Debug route to see what persons are being extracted"""
    if 'id' not in session or session['id'] not in temp_storage:
        return jsonify({'error': 'No data available'})
    
    df = temp_storage[session['id']]['df']
    
    debug_info = []
    persons = set()
    
    for idx, detail in enumerate(df['Details']):
        if pd.isna(detail) or detail == '':
            continue
            
        detail_str = str(detail).upper()
        extracted = []
        
        # Extract UPI IDs (common pattern: name@bank)
        upi_matches = re.findall(r'([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})', detail_str)
        for upi in upi_matches:
            name_part = upi.split('@')[0]
            if len(name_part) > 2:
                persons.add(name_part)
                extracted.append(f"UPI: {name_part}")
        
        # Extract merchant names from UPI transactions
        if 'UPI' in detail_str:
            merchant_match = re.search(r'PAY\s+TO\s+([A-Z\s]+?)(?:\s+UPI|$)', detail_str)
            if merchant_match:
                merchant = merchant_match.group(1).strip()
                if len(merchant) > 2:
                    persons.add(merchant)
                    extracted.append(f"PAY TO: {merchant}")
            
            transfer_match = re.search(r'TRANSFER\s+(?:TO|FROM)\s+([A-Z0-9\s]+?)(?:\s+\d|$)', detail_str)
            if transfer_match:
                account_name = transfer_match.group(1).strip()
                if len(account_name) > 2:
                    persons.add(account_name)
                    extracted.append(f"TRANSFER: {account_name}")
        
        # Extract account numbers and names
        account_match = re.search(r'(\d{10,})', detail_str)
        if account_match:
            account_num = account_match.group(1)
            before_account = detail_str.split(account_num)[0].strip()
            if len(before_account) > 3:
                persons.add(before_account)
                extracted.append(f"BEFORE ACCOUNT: {before_account}")
        
        # Extract common merchant patterns
        if 'HDFC' in detail_str:
            hdfc_match = re.search(r'([A-Z\s]+?)\s*HDFC', detail_str)
            if hdfc_match:
                merchant = hdfc_match.group(1).strip()
                if len(merchant) > 2:
                    persons.add(merchant)
                    extracted.append(f"HDFC: {merchant}")
        
        if 'SBI' in detail_str:
            sbi_match = re.search(r'([A-Z\s]+?)\s*SBI', detail_str)
            if sbi_match:
                merchant = sbi_match.group(1).strip()
                if len(merchant) > 2:
                    persons.add(merchant)
                    extracted.append(f"SBI: {merchant}")
        
        # Extract names from common patterns
        if 'PAY' in detail_str:
            pay_match = re.search(r'PAY\s+([A-Z\s]+?)(?:\s+\d|$)', detail_str)
            if pay_match:
                payee = pay_match.group(1).strip()
                if len(payee) > 2:
                    persons.add(payee)
                    extracted.append(f"PAY: {payee}")
        
        if extracted:
            debug_info.append({
                'index': idx,
                'original': detail,
                'extracted': extracted
            })
    
    return jsonify({
        'all_persons': sorted(list(persons)),
        'debug_info': debug_info[:20]  # Show first 20 for debugging
    })

@app.route('/filter_transactions', methods=['POST'])
def filter_transactions():
    """Filter transactions based on criteria"""
    if 'id' not in session or session['id'] not in temp_storage:
        return jsonify({'error': 'No data available'})
    
    data = request.get_json()
    person = data.get('person', 'All')
    credit_range = data.get('credit_range', 'All')
    debit_range = data.get('debit_range', 'All')
    
    df = temp_storage[session['id']]['df'].copy()
    
    # Apply person filter
    if person != 'All':
        df = df[df['Details'].str.contains(person, case=False, na=False)]
    
    # Apply credit amount filter
    if credit_range != 'All':
        range_map = {
            '₹0 - ₹1,000': (0, 1000),
            '₹1,000 - ₹5,000': (1000, 5000),
            '₹5,000 - ₹10,000': (5000, 10000),
            '₹10,000 - ₹50,000': (10000, 50000),
            '₹50,000+': (50000, float('inf'))
        }
        if credit_range in range_map:
            min_val, max_val = range_map[credit_range]
            df['Credit_float'] = df['Credit'].replace('', '0').astype(float)
            if max_val == float('inf'):
                df = df[df['Credit_float'] >= min_val]
            else:
                df = df[(df['Credit_float'] >= min_val) & (df['Credit_float'] < max_val)]
            df = df.drop('Credit_float', axis=1)
    
    # Apply debit amount filter
    if debit_range != 'All':
        range_map = {
            '₹0 - ₹1,000': (0, 1000),
            '₹1,000 - ₹5,000': (1000, 5000),
            '₹5,000 - ₹10,000': (5000, 10000),
            '₹10,000 - ₹50,000': (10000, 50000),
            '₹50,000+': (50000, float('inf'))
        }
        if debit_range in range_map:
            min_val, max_val = range_map[debit_range]
            df['Debit_float'] = df['Debit'].replace('', '0').astype(float)
            if max_val == float('inf'):
                df = df[df['Debit_float'] >= min_val]
            else:
                df = df[(df['Debit_float'] >= min_val) & (df['Debit_float'] < max_val)]
            df = df.drop('Debit_float', axis=1)
    
    # Calculate filtered summary
    debit_sum = df['Debit'].replace('', '0').astype(float).sum()
    credit_sum = df['Credit'].replace('', '0').astype(float).sum()
    balance = credit_sum - debit_sum
    
    return jsonify({
        'filtered_data': df.to_dict('records'),
        'summary': {
            'transaction_count': len(df),
            'debit': debit_sum,
            'credit': credit_sum,
            'balance': balance
        }
    })

@app.route('/export_filtered/xlsx')
def export_filtered_xlsx():
    """Export filtered transactions to Excel"""
    if 'id' not in session or session['id'] not in temp_storage:
        return redirect(url_for('index'))
    
    # Get filter parameters from query string
    person = request.args.get('person', 'All')
    credit_range = request.args.get('credit_range', 'All')
    debit_range = request.args.get('debit_range', 'All')
    
    df = temp_storage[session['id']]['df'].copy()
    filename = temp_storage[session['id']]['filename']
    base_name = os.path.splitext(filename)[0]
    
    # Apply same filters as in filter_transactions
    if person != 'All':
        df = df[df['Details'].str.contains(person, case=False, na=False)]
    
    if credit_range != 'All':
        range_map = {
            '₹0 - ₹1,000': (0, 1000),
            '₹1,000 - ₹5,000': (1000, 5000),
            '₹5,000 - ₹10,000': (5000, 10000),
            '₹10,000 - ₹50,000': (10000, 50000),
            '₹50,000+': (50000, float('inf'))
        }
        if credit_range in range_map:
            min_val, max_val = range_map[credit_range]
            df['Credit_float'] = df['Credit'].replace('', '0').astype(float)
            if max_val == float('inf'):
                df = df[df['Credit_float'] >= min_val]
            else:
                df = df[(df['Credit_float'] >= min_val) & (df['Credit_float'] < max_val)]
            df = df.drop('Credit_float', axis=1)
    
    if debit_range != 'All':
        range_map = {
            '₹0 - ₹1,000': (0, 1000),
            '₹1,000 - ₹5,000': (1000, 5000),
            '₹5,000 - ₹10,000': (5000, 10000),
            '₹10,000 - ₹50,000': (10000, 50000),
            '₹50,000+': (50000, float('inf'))
        }
        if debit_range in range_map:
            min_val, max_val = range_map[debit_range]
            df['Debit_float'] = df['Debit'].replace('', '0').astype(float)
            if max_val == float('inf'):
                df = df[df['Debit_float'] >= min_val]
            else:
                df = df[(df['Debit_float'] >= min_val) & (df['Debit_float'] < max_val)]
            df = df.drop('Debit_float', axis=1)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Transactions', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Transactions']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    output.seek(0)
    
    # Create filename with filter info
    filter_parts = []
    if person != 'All':
        filter_parts.append(f"Person_{person}")
    if credit_range != 'All':
        filter_parts.append(f"Credit_{credit_range.replace(' ', '_')}")
    if debit_range != 'All':
        filter_parts.append(f"Debit_{debit_range.replace(' ', '_')}")
    
    if filter_parts:
        download_name = f'{base_name}_filtered_{"_".join(filter_parts)}.xlsx'
    else:
        download_name = f'{base_name}_all_transactions.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@app.route('/export_filtered/csv')
def export_filtered_csv():
    """Export filtered transactions to CSV"""
    if 'id' not in session or session['id'] not in temp_storage:
        return redirect(url_for('index'))
    
    # Get filter parameters from query string
    person = request.args.get('person', 'All')
    credit_range = request.args.get('credit_range', 'All')
    debit_range = request.args.get('debit_range', 'All')
    
    df = temp_storage[session['id']]['df'].copy()
    filename = temp_storage[session['id']]['filename']
    base_name = os.path.splitext(filename)[0]
    
    # Apply same filters as in filter_transactions
    if person != 'All':
        df = df[df['Details'].str.contains(person, case=False, na=False)]
    
    if credit_range != 'All':
        range_map = {
            '₹0 - ₹1,000': (0, 1000),
            '₹1,000 - ₹5,000': (1000, 5000),
            '₹5,000 - ₹10,000': (5000, 10000),
            '₹10,000 - ₹50,000': (10000, 50000),
            '₹50,000+': (50000, float('inf'))
        }
        if credit_range in range_map:
            min_val, max_val = range_map[credit_range]
            df['Credit_float'] = df['Credit'].replace('', '0').astype(float)
            if max_val == float('inf'):
                df = df[df['Credit_float'] >= min_val]
            else:
                df = df[(df['Credit_float'] >= min_val) & (df['Credit_float'] < max_val)]
            df = df.drop('Credit_float', axis=1)
    
    if debit_range != 'All':
        range_map = {
            '₹0 - ₹1,000': (0, 1000),
            '₹1,000 - ₹5,000': (1000, 5000),
            '₹5,000 - ₹10,000': (5000, 10000),
            '₹10,000 - ₹50,000': (10000, 50000),
            '₹50,000+': (50000, float('inf'))
        }
        if debit_range in range_map:
            min_val, max_val = range_map[debit_range]
            df['Debit_float'] = df['Debit'].replace('', '0').astype(float)
            if max_val == float('inf'):
                df = df[df['Debit_float'] >= min_val]
            else:
                df = df[(df['Debit_float'] >= min_val) & (df['Debit_float'] < max_val)]
            df = df.drop('Debit_float', axis=1)
    
    # Create CSV in memory
    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)
    
    # Create filename with filter info
    filter_parts = []
    if person != 'All':
        filter_parts.append(f"Person_{person}")
    if credit_range != 'All':
        filter_parts.append(f"Credit_{credit_range.replace(' ', '_')}")
    if debit_range != 'All':
        filter_parts.append(f"Debit_{debit_range.replace(' ', '_')}")
    
    if filter_parts:
        download_name = f'{base_name}_filtered_{"_".join(filter_parts)}.csv'
    else:
        download_name = f'{base_name}_all_transactions.csv'
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=download_name
    )

@app.route('/clear_session')
def clear_session():
    sid = session.get('id')
    if sid and sid in temp_storage:
        del temp_storage[sid]
    session.clear()
    return redirect('/')

@app.errorhandler(404)
def not_found(error):
    return render_template('error.html', error="Page not found"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', error="Internal server error"), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)